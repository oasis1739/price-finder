import asyncio
import io
import os
import re
from typing import List, Optional
import openpyxl
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, RedirectResponse
from pydantic import BaseModel

from scrapers import ALL_SCRAPERS, SCRAPER_MAP
from scrapers.base import SearchResult

app = FastAPI(title="최저가 검색기")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")


@app.middleware("http")
async def prevent_html_cache(request, call_next):
    response = await call_next(request)
    if request.url.path == "/" or request.url.path.endswith(".html"):
        response.headers["Cache-Control"] = "no-store"
    return response


class SearchRequest(BaseModel):
    product_number: str = ""
    product_name: str = ""
    sites: Optional[List[str]] = None


class ProductItem(BaseModel):
    product_number: str = ""
    product_name: str = ""


class BulkSearchRequest(BaseModel):
    items: List[ProductItem]
    sites: Optional[List[str]] = None


def normalize_product_number(raw: str) -> str:
    raw = raw.strip()
    match = re.search(r"[A-Za-z0-9][A-Za-z0-9\-_]+[A-Za-z0-9]", raw)
    return match.group(0) if match else raw


async def parse_excel_items(file: UploadFile) -> list[ProductItem]:
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 파일을 읽을 수 없습니다")

    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    num_col = next((i for i, h in enumerate(headers) if "품번" in h or "코드" in h or "번호" in h), None)
    name_col = next((i for i, h in enumerate(headers) if "상품명" in h or "상품" in h or "이름" in h), None)

    if num_col is None and name_col is None:
        num_col, name_col = 0, 1

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pnum = str(row[num_col] or "").strip() if num_col is not None and num_col < len(row) else ""
        pname = str(row[name_col] or "").strip() if name_col is not None and name_col < len(row) else ""
        if pnum or pname:
            items.append(ProductItem(product_number=pnum, product_name=pname))

    if not items:
        raise HTTPException(status_code=400, detail="엑셀에서 상품 데이터를 찾을 수 없습니다")

    return items


async def run_search(product_number: str, product_name: str, sites: Optional[list[str]]) -> list[dict]:
    pnum = normalize_product_number(product_number) if product_number else ""

    if sites:
        scrapers = [SCRAPER_MAP[s]() for s in sites if s in SCRAPER_MAP]
    else:
        scrapers = [cls() for cls in ALL_SCRAPERS]

    tasks = [s.search(pnum, product_name) for s in scrapers]
    all_results: list[SearchResult] = []

    results_per_site = await asyncio.gather(*tasks, return_exceptions=True)
    for res in results_per_site:
        if isinstance(res, Exception):
            continue
        all_results.extend(res)

    all_results.sort(key=lambda x: x.total_price)

    return [
        {
            "site": r.site,
            "product_name": r.product_name,
            "product_number": r.product_number,
            "price": r.price,
            "shipping_fee": r.shipping_fee,
            "total_price": r.total_price,
            "url": r.url,
            "image_url": r.image_url,
            "matched_by": r.matched_by,
        }
        for r in all_results
    ]


@app.get("/api/sites")
def get_sites():
    has_naver = bool(os.getenv("NAVER_CLIENT_ID") and os.getenv("NAVER_CLIENT_SECRET"))
    return [
        {
            "name": cls.site_name,
            "needs_api": cls.site_name in ("네이버쇼핑", "GS샵", "퀸잇"),
            "api_available": has_naver if cls.site_name in ("네이버쇼핑", "GS샵", "퀸잇") else True,
        }
        for cls in ALL_SCRAPERS
    ]


@app.get("/api/status")
def get_status():
    return {
        "naver_api": bool(os.getenv("NAVER_CLIENT_ID") and os.getenv("NAVER_CLIENT_SECRET")),
    }


@app.post("/api/search")
async def search(req: SearchRequest):
    if not req.product_number and not req.product_name:
        raise HTTPException(status_code=400, detail="품번 또는 상품명을 입력해주세요")
    results = await run_search(req.product_number, req.product_name, req.sites)
    return {"results": results, "count": len(results)}


@app.post("/api/search/bulk")
async def search_bulk(req: BulkSearchRequest):
    if not req.items:
        raise HTTPException(status_code=400, detail="검색할 상품이 없습니다")

    all_output = []
    for item in req.items:
        results = await run_search(item.product_number, item.product_name, req.sites)
        all_output.append({
            "input_number": item.product_number,
            "input_name": item.product_name,
            "results": results,
        })

    return {"items": all_output}


@app.post("/api/search/excel")
async def search_from_excel(file: UploadFile = File(...), sites: Optional[str] = None):
    items = await parse_excel_items(file)
    site_list = sites.split(",") if sites else None

    all_output = []
    for item in items:
        results = await run_search(item.product_number, item.product_name, site_list)
        all_output.append({
            "input_number": item.product_number,
            "input_name": item.product_name,
            "results": results,
        })

    return {"items": all_output}


@app.post("/api/excel/items")
async def get_items_from_excel(file: UploadFile = File(...)):
    items = await parse_excel_items(file)
    return {
        "items": [
            {"product_number": item.product_number, "product_name": item.product_name}
            for item in items
        ],
        "count": len(items),
    }


@app.get("/api/redirect/fashionplus/{goods_id}")
def redirect_fashionplus(goods_id: str):
    """패션플러스 상품으로 리다이렉트 (403 방지)"""
    return RedirectResponse(
        url=f"https://www.fashionplus.co.kr/goods/detail/{goods_id}",
        status_code=301
    )


@app.get("/api/export/excel")
async def export_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품목록"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.append(["품번", "상품명"])
    ws.append(["WBTM509", "썸머 쿨 하프소매 볼륨 니트"])
    ws.append(["AB-1234", "여름 반팔 니트 티셔츠"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template.xlsx"},
    )


# 정적 파일
if os.path.exists(FRONTEND_DIR):
    app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="static")
